"""Excel parser for OpenNEM"""
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from opennem.utils.mime import mime_from_content
from opennem.utils.xls import convert_to_xlxs


def detect_sheat_data_dimensions(sheet: Worksheet) -> tuple[int, int]:
    """Calculates how many rows and columns to skip on a sheet"""
    min_col = 0
    min_row = 0

    for col in sheet.columns:
        _col_values = [i.value for i in col if i.value]

        if len(_col_values) == 0:
            min_col += 1
        else:
            break

    for row in sheet.rows:
        _row_values = [i.value for i in row if i.value]

        if len(_row_values) < sheet.max_column:
            min_row += 1
        else:
            break

    return min_col, min_row


def parse_workbook_sheet(sheet: Worksheet) -> list[dict]:
    min_col, min_row = detect_sheat_data_dimensions(sheet)

    print(f"{sheet.title} has min_col {min_col} and min_row {min_row}")

    return []


def parse_workbook(
    fh: BinaryIO,
    convert_xls: bool = True,
    worksheet: str | None = None,
) -> Workbook | Worksheet:
    """Parse an excel file (with conversion) into a dict of lists for each sheet"""

    fh.seek(0)
    fh_mime = mime_from_content(fh)
    wb: Workbook | None = None

    if fh_mime in ["application/CDFV2"]:
        if convert_xls:
            wb = convert_to_xlxs(fh)
    else:
        wb = load_workbook(fh)

    if not wb:
        raise Exception("Could not parse workbook")

    if worksheet:
        if worksheet not in wb:
            raise Exception(f"Could not find worksheet {worksheet} in workbook")

        return wb[worksheet]

    return wb
