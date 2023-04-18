import logging
from io import BytesIO
from itertools import groupby
from pathlib import Path
from typing import IO, Any

from openpyxl import load_workbook

from opennem.core.dispatch_type import parse_dispatch_type
from opennem.core.facility.fueltechs import parse_facility_fueltech
from opennem.core.facilitystatus import parse_facility_status
from opennem.core.fueltechs import lookup_fueltech
from opennem.core.loader import PROJECT_DATA_PATH, load_data
from opennem.core.normalizers import clean_capacity, normalize_duid, station_name_cleaner
from opennem.core.parsers.excel import parse_workbook
from opennem.core.stations.station_code_from_duids import station_code_from_duids
from opennem.core.stations.station_name_code_map import station_name_code_map
from opennem.core.unit_codes import get_unit_code
from opennem.core.unit_parser import parse_unit_duid
from opennem.schema.aemo.rel import AEMORelGeneratorRecord
from opennem.schema.stations import StationSet

logger = logging.getLogger("opennem.parsers.aemo.rel")

participant_keys = ["name", "abn"]

AEMO_REL_GENERATOR_SHEETNAME = "Generators and Scheduled Loads"

FACILITY_KEYS = [
    "participant",
    "station_name",
    "region",
    "dispatch_type",
    "category",
    "classification",
    "fuel_source_primary",
    "fuel_source_descriptor",
    "tech_primary",
    "tech_primary_descriptor",
    "unit_no",
    "unit_size",
    "aggreagation",
    "duid",
    "reg_cap",
    "max_cap",
    "max_roc",
]


def lookup_station_code(duids: list[str], station_name: str, station_code_map: dict) -> str | None:
    station_code = None
    station_name = station_name.strip()

    for facility_duid in duids:
        if facility_duid in station_code_map:
            station_code = station_code_map[facility_duid]
            continue

    if not station_code:
        station_code = station_name_code_map(station_name)

    if not station_code:
        station_code = station_code_from_duids(duids)

    if not station_code:
        print("Coult not get a station code for {}: {}".format(station_name, ",".join(duids)))

    return station_code


def rel_grouper(records: list[dict], station_code_map: dict) -> dict[str, Any]:
    records_parsed = []

    for _id, i in enumerate(records, start=2000):
        name = station_name_cleaner(i["station_name"])
        duid = normalize_duid(i["duid"])

        if not duid:
            raise Exception(f"No duid for {name}")

        unit = parse_unit_duid(i["unit_no"], duid)
        fueltech = lookup_fueltech(
            i["fuel_source_primary"],
            i["fuel_source_descriptor"],
            i["tech_primary"],
            i["tech_primary_descriptor"],
            i["dispatch_type"],
        )
        station_code = lookup_station_code([duid], i["station_name"], station_code_map)

        records_parsed.append(
            {
                "name": name,
                "code": duid,
                "status": parse_facility_status("operating"),
                "station_code": station_code,
                "network_region": i["region"].strip(),
                "network_name": i["station_name"].strip(),
                "unit_size": clean_capacity(i["unit_size"]),
                "unit_code": get_unit_code(unit, duid, name),
                "dispatch_type": parse_dispatch_type(i["dispatch_type"]),
                "fueltech": parse_facility_fueltech(fueltech),
                "capacity_registered": clean_capacity(i["reg_cap"]),
                "capacity_maximum": clean_capacity(i["max_cap"]),
            }
        )

    grouped_records: dict[str, Any] = {}

    for key, v in groupby(records_parsed, key=lambda v: v["station_code"]):
        station_code = str(key)

        # key = k[1
        if key not in grouped_records:
            grouped_records[station_code] = []

        grouped_records[station_code] += list(v)

    coded_records: dict[str, Any] = {}
    _id = 2000

    for station_code, rel in grouped_records.items():
        station_name = rel[0]["network_name"]

        if station_code in coded_records:
            raise Exception(f"Code conflict: {station_code}. {station_name} {coded_records[station_code]}")

        if not station_code:
            raise Exception(f"Unmapped station: {rel}")

        coded_records[station_code] = {
            "name": station_name_cleaner(station_name),
            "network_name": station_name,
            "code": station_code,
            "id": _id,
            "facilities": rel,
        }

        _id += 1

    return coded_records


def load_rel() -> list[dict[str, Any]]:
    aemo_path: Path = Path(PROJECT_DATA_PATH) / "aemo" / "nem_rel_202108.xls"

    if not aemo_path.is_file():
        raise Exception(f"Not found: {aemo_path}")

    wb = load_workbook(aemo_path, data_only=True)

    if "Generators and Scheduled Loads" not in wb:
        raise Exception("Invalid REL formatted spreadsheet")

    generator_ws = wb["Generators and Scheduled Loads"]
    # participant_ws = wb.get_sheet_by_name("Registered Participants")

    generators = []

    for row in generator_ws.iter_rows(min_row=2, values_only=True):
        generators.append(dict(zip(FACILITY_KEYS, list(row[: len(FACILITY_KEYS)]), strict=True)))

    return generators


def rel_import() -> StationSet:
    mms_duid_station_map = load_data("mms_duid_station_map.json", True)

    nem_rel = load_rel()
    nem_rel_dict = rel_grouper(nem_rel, mms_duid_station_map)

    rel = StationSet()

    for s in nem_rel_dict.values():
        rel.add_dict(s)

    return rel


def rel_export() -> None:
    nem_rel = rel_import()

    with open("data/rel.json", "w") as fh:
        fh.write(nem_rel.json(indent=4))

    logger.info(f"Wrote {nem_rel.length} records")


# new


def parse_aemo_rel_generators(filename: str | Path | BytesIO) -> list[AEMORelGeneratorRecord]:
    fh: BytesIO | None = None

    if isinstance(filename, str):
        filename = Path(filename)

    if isinstance(filename, Path):
        if not filename.is_file():
            raise Exception(f"Could not open specified rel file {filename}")

        fh = open(str(filename), "rb")

    if isinstance(filename, IO):
        fh = filename

    if not fh or not isinstance(fh, BytesIO):
        raise Exception("Could not open file")

    generators_ws = parse_workbook(fh, worksheet=AEMO_REL_GENERATOR_SHEETNAME)

    generator_records = []

    for row in generators_ws.iter_rows(min_row=2, values_only=True):
        _row_dict = dict(zip(FACILITY_KEYS, list(row[: len(FACILITY_KEYS)]), strict=True))

        _row_schema = AEMORelGeneratorRecord(**_row_dict)
        generator_records.append(_row_schema)

    # s = StationSet()

    return generator_records


if __name__ == "__main__":
    aemo_path: Path = Path(PROJECT_DATA_PATH) / "aemo" / "nem_rel_202108.xls"
    r = parse_aemo_rel_generators("opennem/data/aemo/nem_rel_202108.xls")

    from pprint import pprint

    pprint(r)

    pprint([i.fueltech_id for i in r])
