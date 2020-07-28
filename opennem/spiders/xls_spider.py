from io import BytesIO
from itertools import groupby

import scrapy
from openpyxl import load_workbook


class XLSSpider(scrapy.Spider):
    def start_requests(self):
        if not hasattr(self, "url"):
            raise Exception("Require URL param")

        yield scrapy.Request(self.url)

    def parse(self, response):
        wb = load_workbook(BytesIO(response.body), data_only=True)

        ws = wb.get_sheet_by_name("ExistingGeneration&NewDevs")

        for row in ws.iter_rows(min_row=3, values_only=True):

            # pick out the columns we want
            # lots of hidden columns in the sheet
            row_collapsed = row[0:8] + (row[11],) + row[12:19] + row[22:]

            return_dict = dict(zip(self.keys, list(row_collapsed)))

            if return_dict is None:
                raise Exception("Failed on row: {}".format(row))

            yield return_dict
