import contextlib
from io import BytesIO

with contextlib.suppress(ImportError):
    from openpyxl import load_workbook


# class AEMOGeneralInformationCurrentSpider(scrapy.Spider):
class Spider:
    """
    Extracts station and unit data from the AEMO general information spreadsheet


    """

    # name = "au.aemo.current.general_information"

    start_urls = [
        # "https://aemo.com.au/-/media/files/electricity/nem/planning_and_forecasting/generation_information/nem-generation-information-april-2020.xlsx?la=en",
        # "https://data.opennem.org.au/v3/data/NEM+Generation+Information+April+2020.xlsx",
        "https://data.opennem.org.au/v3/data/NEM+Generation+Information+July+2020.xlsx"
    ]

    keys = [
        "Region",
        "Status",
        "station_name",
        "Owner",
        "TechType",
        "FuelType",
        "duid",
        "Units",
        "unit_capacity_lower",
        "unit_capacity",
        "UpperCapacity",
        "NameCapacity",
        "StorageCapacity",
        "UnitStatus",
        "DispatchType",
        "UseDate",
        "ClosureDateExpected",
        "ClosureDate",
        "SurveyID",
        "SurveyVersion",
        "SurveyEffective",
    ]

    def parse(self, response):
        wb = load_workbook(BytesIO(response.body), data_only=True)

        ws = wb.get_sheet_by_name("ExistingGeneration&NewDevs")

        records = []

        for row in ws.iter_rows(min_row=3, values_only=True):

            # pick out the columns we want
            # lots of hidden columns in the sheet
            row_collapsed = row[:10] + (row[11],) + row[12:19] + row[21:]

            return_dict = dict(zip(self.keys, list(row_collapsed)))

            if return_dict is None:
                raise Exception("Failed on row: {}".format(row))

            records.append(return_dict)

        yield {"records": records}
