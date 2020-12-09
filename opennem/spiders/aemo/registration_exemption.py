from io import BytesIO

import scrapy
from openpyxl import load_workbook

from opennem.pipelines.aemo.registration_exemption import (
    RegistrationExemptionGrouperPipeline,
    RegistrationExemptionStorePipeline,
)


class AEMORegistrationExemptionListSpider(scrapy.Spider):
    """
    Crawls the current AEMO registration and exemption spreadsheet
    and extracts participants and generators

    """

    # name = "au.aemo.current.registration_exemption"

    start_urls = [
        "https://data.opennem.org.au/v3/data/NEM+Registration+and+Exemption+List+July.xlsx"
    ]

    participant_keys = ["name", "abn"]

    generator_keys = [
        "participant",
        "station_name",
        "region",
        "dispatch_type",
        "category",
        "classification",
        "fuel_source_primary",
        "fuel_source_descriptor",
        "tech_primary",
        "tech_primary_descriptor",
        "unit_no",
        "unit_size",
        "aggreagation",
        "duid",
        "reg_cap",
        "max_cap",
        "max_roc",
    ]

    pipelines_extra = set(
        [
            RegistrationExemptionGrouperPipeline,
            RegistrationExemptionStorePipeline,
        ]
    )

    def parse(self, response):

        wb = load_workbook(BytesIO(response.body), data_only=True)

        generator_ws = wb.get_sheet_by_name("Generators and Scheduled Loads")
        participant_ws = wb.get_sheet_by_name("Registered Participants")

        generators = []
        participants = []

        for row in generator_ws.iter_rows(min_row=2, values_only=True):
            generators.append(
                dict(
                    zip(
                        self.generator_keys,
                        list(row[0 : len(self.generator_keys)]),
                    )
                )
            )

        for row in participant_ws.iter_rows(min_row=3, values_only=True):

            participants.append(
                dict(
                    zip(
                        self.participant_keys,
                        list(row[0 : len(self.participant_keys)]),
                    )
                )
            )

        participants = sorted(participants, key=lambda k: k["name"])

        yield {
            "generators": generators,
            "participants": participants,
        }
