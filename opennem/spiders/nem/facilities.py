from io import BytesIO

import scrapy
from openpyxl import load_workbook

from opennem.pipelines.nem.facilities import (
    NemStoreGI,
    NemStoreMMS,
    NemStoreREL,
)


class NemFacilitySpider(scrapy.Spider):
    name = "au.nem.facilities.gi"

    start_urls = [
        # "https://aemo.com.au/-/media/files/electricity/nem/planning_and_forecasting/generation_information/nem-generation-information-april-2020.xlsx?la=en",
        # "https://data.opennem.org.au/v3/data/NEM+Generation+Information+April+2020.xlsx",
        "https://data.opennem.org.au/v3/data/NEM+Generation+Information+July+2020.xlsx"
    ]

    keys = [
        "Region",
        "Status",
        "Name",
        "Owner",
        "TechType",
        "FuelType",
        "DUID",
        "Units",
        "UpperCapacity",
        "NameCapacity",
        "StorageCapacity",
        "UnitStatus",
        "DispatchType",
        "UseDate",
        "ClosureDateExpected",
        "ClosureDate",
        "SurveyVersion",
        "SurveyEffective",
    ]

    pipelines_extra = set([NemStoreGI,])

    def parse(self, response):
        wb = load_workbook(BytesIO(response.body), data_only=True)

        ws = wb.get_sheet_by_name("ExistingGeneration&NewDevs")

        for row in ws.iter_rows(min_row=3, values_only=True):

            # pick out the columns we want
            # lots of hidden columns in the sheet
            row_collapsed = row[0:8] + (row[11],) + row[12:19] + row[22:]

            return_dict = dict(zip(self.keys, list(row_collapsed)))

            if return_dict is None:
                raise Exception("Failed on row: {}".format(row))

            yield return_dict


class NemParticipantSpider(scrapy.Spider):
    name = "au.nem.facilities.rel"

    start_urls = [
        "https://data.opennem.org.au/v3/data/NEM+Registration+and+Exemption+List.xlsx"
    ]

    participant_keys = ["name", "abn"]

    generator_keys = [
        "participant",
        "station_name",
        "region",
        "dispatch_type",
        "category",
        "classification",
        "fuel_source_primary",
        "fuel_source_descriptor",
        "tech_primary",
        "tech_primary_descriptor",
        "unit_no",
        "unit_size",
        "aggreagation",
        "duid",
        "reg_cap",
        "max_cap",
        "max_roc",
    ]

    pipelines_extra = set([NemStoreREL,])

    def parse(self, response):

        wb = load_workbook(BytesIO(response.body), data_only=True)

        generator_ws = wb.get_sheet_by_name("Generators and Scheduled Loads")
        participant_ws = wb.get_sheet_by_name("Registered Participants")

        generators = []
        participants = []

        for row in generator_ws.iter_rows(min_row=2, values_only=True):
            generators.append(
                dict(
                    zip(
                        self.generator_keys,
                        list(row[0 : len(self.generator_keys)]),
                    )
                )
            )

        for row in participant_ws.iter_rows(min_row=3, values_only=True):

            participants.append(
                dict(
                    zip(
                        self.participant_keys,
                        list(row[0 : len(self.participant_keys)]),
                    )
                )
            )

        yield {"generators": generators, "participants": participants}


class NemFacilityMMSSpider(scrapy.Spider):
    name = "au.nem.facility.mms"

    start_urls = [
        "http://nemweb.com.au/Data_Archive/Wholesale_Electricity/MMSDM/2020/MMSDM_2020_06/MMSDM_Historical_Data_SQLLoader/DATA/PUBLIC_DVD_DUDETAILSUMMARY_202006010000.zip"
    ]

    pipelines_extra = set([NemStoreREL,])

    def parse(self, response):
        yield response
