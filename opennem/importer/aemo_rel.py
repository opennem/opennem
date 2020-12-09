import json
import logging
from itertools import groupby
from typing import List, Optional

from openpyxl import load_workbook

from opennem.core.dispatch_type import parse_dispatch_type
from opennem.core.facility.fueltechs import parse_facility_fueltech
from opennem.core.facilitystatus import parse_facility_status
from opennem.core.fueltechs import lookup_fueltech
from opennem.core.loader import PROJECT_DATA_PATH, load_data
from opennem.core.normalizers import (
    clean_capacity,
    normalize_duid,
    station_name_cleaner,
)
from opennem.core.stations.station_code_from_duids import (
    station_code_from_duids,
)
from opennem.core.stations.station_name_code_map import station_name_code_map
from opennem.core.unit_codes import get_unit_code
from opennem.core.unit_parser import parse_unit_duid
from opennem.exporter.encoders import OpenNEMJSONEncoder
from opennem.schema.stations import StationSet

logger = logging.getLogger("opennem.importer.mms")

participant_keys = ["name", "abn"]

FACILITY_KEYS = [
    "participant",
    "station_name",
    "region",
    "dispatch_type",
    "category",
    "classification",
    "fuel_source_primary",
    "fuel_source_descriptor",
    "tech_primary",
    "tech_primary_descriptor",
    "unit_no",
    "unit_size",
    "aggreagation",
    "duid",
    "reg_cap",
    "max_cap",
    "max_roc",
]


def lookup_station_code(
    duids: List[str], station_name: str, station_code_map: dict
) -> Optional[str]:

    station_code = None
    station_name = station_name.strip()

    for facility_duid in duids:
        if facility_duid in station_code_map:
            station_code = station_code_map[facility_duid]
            continue

    if not station_code:
        station_code = station_name_code_map(station_name)

    if not station_code:
        station_code = station_code_from_duids(duids)

    if not station_code:
        print(
            "Coult not get a station code for {}: {}".format(
                station_name, ",".join(duids)
            )
        )

    return station_code


def rel_grouper(records, station_code_map):
    records_parsed = []

    for _id, i in enumerate(records, start=2000):
        name = station_name_cleaner(i["station_name"])
        duid = normalize_duid(i["duid"])
        unit = parse_unit_duid(i["unit_no"], duid)
        fueltech = lookup_fueltech(
            i["fuel_source_primary"],
            i["fuel_source_descriptor"],
            i["tech_primary"],
            i["tech_primary_descriptor"],
            i["dispatch_type"],
        )
        station_code = lookup_station_code(
            [duid], i["station_name"], station_code_map
        )

        records_parsed.append(
            {
                "name": name,
                "code": duid,
                "status": parse_facility_status("operating"),
                "station_code": station_code,
                "network_region": i["region"].strip(),
                "network_name": i["station_name"].strip(),
                "unit_size": clean_capacity(i["unit_size"]),
                "unit_code": get_unit_code(unit, duid, name),
                "dispatch_type": parse_dispatch_type(i["dispatch_type"]),
                "fueltech": parse_facility_fueltech(fueltech),
                "capacity_registered": clean_capacity(i["reg_cap"]),
                "capacity_maximum": clean_capacity(i["max_cap"]),
            }
        )

    grouped_records = {}

    for key, v in groupby(records_parsed, key=lambda v: v["station_code"]):

        # key = k[1
        if key not in grouped_records:
            grouped_records[key] = []

        grouped_records[key] += list(v)

    coded_records = {}
    _id = 2000

    for station_code, rel in grouped_records.items():
        station_name = rel[0]["network_name"]

        if station_code in coded_records:
            raise Exception(
                "Code conflict: {}. {} {}".format(
                    station_code, station_name, coded_records[station_code]
                )
            )

        if not station_code:
            raise Exception("Unmapped station: {}".format(rel))

        coded_records[station_code] = {
            "name": station_name_cleaner(station_name),
            "network_name": station_name,
            "code": station_code,
            "id": _id,
            "facilities": rel,
        }

        _id += 1

    return coded_records


def load_rel():
    aemo_path = PROJECT_DATA_PATH / "aemo" / "nem_rel_202006.xlsx"

    if not aemo_path.is_file():
        raise Exception("Not found: {}".format(aemo_path))

    wb = load_workbook(aemo_path, data_only=True)

    generator_ws = wb["Generators and Scheduled Loads"]
    # participant_ws = wb.get_sheet_by_name("Registered Participants")

    generators = []

    for row in generator_ws.iter_rows(min_row=2, values_only=True):
        generators.append(
            dict(
                zip(
                    FACILITY_KEYS,
                    list(row[: len(FACILITY_KEYS)]),
                )
            )
        )

    return generators


def rel_import():
    mms_duid_station_map = load_data("mms_duid_station_map.json", True)

    nem_rel = load_rel()
    nem_rel = rel_grouper(nem_rel, mms_duid_station_map)

    rel = StationSet()

    for s in nem_rel.values():
        rel.add_dict(s)

    return rel


def rel_export():
    nem_rel = rel_import()

    with open("data/rel.json", "w") as fh:
        fh.write(nem_rel.json(indent=4))

    logger.info("Wrote {} records".format(nem_rel.length))


if __name__ == "__main__":
    rel_export()
