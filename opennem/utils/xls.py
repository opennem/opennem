"""
OpenNEM Excel File Utils

@NOTE https://github.com/snoopyjc/xls2xlsx/blob/master/xls2xlsx/xls2xlsx.py
"""

_HAVE_XLRD = False

import logging  # noqa: E402
from typing import BinaryIO, Optional  # noqa: E402

try:
    import xlrd

    _HAVE_XLRD = True
except ImportError:
    pass

from openpyxl.workbook.workbook import Workbook  # noqa: E402

logger = logging.getLogger("opennem.utils.xls")


def convert_to_xlxs(fh: BinaryIO) -> Optional[Workbook]:
    """Convert old workbook formats xls into xlxs"""
    if not _HAVE_XLRD:
        logger.error("xlrd module not installed. Cannot convert XLS file.")
        return None

    file_content = fh.read()

    xlsBook = xlrd.open_workbook(file_contents=file_content)
    workbook = Workbook()

    for i in range(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in range(0, xlsSheet.nrows):
            for col in range(0, xlsSheet.ncols):
                sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)

    return workbook
