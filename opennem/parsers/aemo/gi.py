"""
OpenNEM NEM General Information Parser

This module parses the general information Excel files that are released
by AEMO (almost) every month. The downloads can be found on this page:


<https://www.aemo.com.au/energy-systems/electricity/national-electricity-market-nem/
    nem-forecasting-and-planning/forecasting-and-planning-data/generation-information>

In opennem.spiders.aemo.monitoring there is a spider that finds the new files
and they will be passed into this file to be output into a StationSet schema,
ready for manipulation, export into different formats - or, for our primary
purpose - to import new facilities and updates into the OpenNEM database.
"""

import csv
import logging
from datetime import date, datetime
from pathlib import Path
from tempfile import mkdtemp
from typing import Any

from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

from opennem import settings
from opennem.core.normalizers import station_name_cleaner
from opennem.db import SessionLocal
from opennem.db.models.opennem import AEMOFacilityData
from opennem.parsers.aemo.normalisers import aemo_gi_fueltech_to_fueltech, aemo_gi_status_map
from opennem.parsers.aemo.schemas import AEMODataSource, AEMOGIRecord, AEMOSourceSet
from opennem.utils.dates import get_today_opennem
from opennem.utils.http import download_file

# from rich.console import Console
# from rich.prompt import Prompt


logger = logging.getLogger("opennem.parsers.aemo.gi")


AEMO_GI_DOWNLOAD_URL = (
    "https://www.aemo.com.au/-/media/files/electricity/nem/planning_and_forecasting/"
    "generation_information/{year}/nem-generation-information-{month}-{year}.xlsx?la=en"
)


class OpennemAEMOGIParserException(Exception):
    """Exception raised in this module"""

    pass


# Maps column keys to names
GI_EXISTING_NEW_GEN_KEYS = {
    "region": "A",
    "AssetType": "B",
    "StationName": "C",
    "Owner": "D",
    "TechType": "E",
    "FuelType": "F",
    "duid": "G",
    "units_no": "H",
    "capacity_registered": "M",
    "StorageCapacity": "N",
    "UnitStatus": "O",
    "DispatchType": "P",
    "UseDate": "Q",
    "ClosureDateExpected": "R",
    "ClosureDate": "S",
    "SurveyID": "V",
    "FuelSummary": "U",
    "SurveyEffective": "Y",
}


def excel_column_to_column_index(excel_column: str) -> int:
    """Takes an excel column like 'A' and converts to column index"""
    return ord(excel_column.upper()) - 64


def parse_aemo_general_information(filename: Path) -> list[AEMOGIRecord]:
    """Primary record parser for GI information. Takes the spreadsheet location from a path,
    parses it and the relevant sheet and returns a list of GI records"""

    wb = load_workbook(str(filename), data_only=True)

    SHEET_KEY = "ExistingGeneration&NewDevs"

    if SHEET_KEY not in wb:
        raise OpennemAEMOGIParserException("Doesn't look like a GI spreadsheet")

    ws = wb[SHEET_KEY]

    records = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        # pick out the columns we want
        # lots of hidden columns in the sheet
        row_collapsed = [row[excel_column_to_column_index(i) - 1] for i in GI_EXISTING_NEW_GEN_KEYS.values()]

        return_dict = dict(zip(GI_EXISTING_NEW_GEN_KEYS, list(row_collapsed)))

        # break at end of data records
        # GI has a blank line before garbage notes
        if row[0] is None:
            break

        if return_dict is None:
            raise OpennemAEMOGIParserException(f"Failed on row: {row}")

        if return_dict["region"] is None or return_dict["region"] not in ["NSW1", "QLD1", "SA1", "TAS1", "VIC1"]:
            continue

        return_dict = {
            **return_dict,
            **{
                "name": station_name_cleaner(return_dict["StationName"]),
                "name_network": return_dict["StationName"],
                "network_region": return_dict["region"].strip().upper(),
                "fueltech_id": aemo_gi_fueltech_to_fueltech(return_dict["FuelSummary"]),
                "status_id": aemo_gi_status_map(return_dict["UnitStatus"]),
                "duid": return_dict["duid"],
                "units_no": return_dict["units_no"],
                "capacity_registered": return_dict["capacity_registered"],
                "closure_year_expected": return_dict["ClosureDateExpected"],
                "unique_id": str(return_dict["SurveyID"]).strip(),
            },
        }

        return_model = AEMOGIRecord(**return_dict)

        records.append(return_model)

    logger.info(f"Parsed {len(records)} records")

    return records


def get_unique_values_for_field(records: list[dict], field_name: str) -> list[Any]:
    return list({i[field_name] for i in records})


def get_aemo_gi_download_url(month: date) -> str:
    """Get the AEMO GI download URL for a given month"""
    return AEMO_GI_DOWNLOAD_URL.format(month=month.strftime("%b").lower(), year=month.strftime("%Y"))


def check_latest_persisted_aemo_source(source_type: str = "gi") -> date | None:
    """Check what the latest month is that we have in the database"""
    with SessionLocal() as session:
        query = (
            session.query(AEMOFacilityData)
            .filter(AEMOFacilityData.aemo_source == source_type)
            .order_by(AEMOFacilityData.source_date.desc())
        )

        if query.count() > 0:
            return query.first()

    return None


def download_latest_aemo_gi_file(num_months_to_check: int = 3) -> AEMOSourceSet:
    """This will download the latest GI file into a local temp directory
    and return the path to it"""

    gi_saved_path: Path | None = None
    dest_dir = Path(mkdtemp(prefix=f"{settings.tmp_file_prefix}"))

    this_month = get_today_opennem()
    months_to_check: list[datetime] = [this_month]

    for i in range(1, num_months_to_check):
        months_to_check.append(this_month - relativedelta(months=i))

    for download_date in months_to_check:
        download_url = get_aemo_gi_download_url(download_date)

        logger.info(f"Downloading AEMO GI file from {download_url}")

        try:
            gi_saved_path = download_file(
                download_url, dest_dir, expect_content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            if gi_saved_path:
                source_set = AEMOSourceSet(
                    aemo_source=AEMODataSource.gi, source_date=download_date, source_url=download_url, local_path=gi_saved_path
                )
                return source_set
        except Exception as e:
            logger.error(f"Failed to download AEMO GI file: {e}")
            continue

    raise OpennemAEMOGIParserException(f"Failed to download AEMO GI file for last {num_months_to_check} months")


def persist_source_set_to_database(source_set: AEMOSourceSet) -> None:
    """Take an AEMOSourceSet and persist it to the database table AEMOFacilityData"""
    pass


def persist_source_set_to_csv(source_set: AEMOSourceSet, filename: str = "aemo_gi.csv") -> None:
    """Take an AEMOSourceSet and persist it to a CSV file"""

    with open(filename, "w") as fh:
        csv_writer = csv.DictWriter(fh, fieldnames=AEMOGIRecord.schema().get("properties").keys())  # type: ignore
        csv_writer.writeheader()

        for record in source_set.records:
            csv_writer.writerow(record.dict())

    logger.info(f"Saved {len(source_set.records)} records to {filename}")


def download_and_parse_aemo_gi_file() -> AEMOSourceSet:
    """This will download the latest GI file and parse it"""
    aemo_source_set = download_latest_aemo_gi_file()

    if not aemo_source_set.local_path:
        raise OpennemAEMOGIParserException("Failed to download AEMO GI file: no local_path")

    aemo_source_set.records = parse_aemo_general_information(aemo_source_set.local_path)

    return aemo_source_set


# debug entrypoint
if __name__ == "__main__":
    aemo_source_set = download_and_parse_aemo_gi_file()
    persist_source_set_to_csv(aemo_source_set)

    for s in aemo_source_set.records:
        print(
            f"{s.name} {s.name_network} {s.network_region} {s.fueltech_id}"
            f" {s.status_id} {s.duid} {s.units_no} {s.capacity_registered} {s.closure_year_expected}"
        )
